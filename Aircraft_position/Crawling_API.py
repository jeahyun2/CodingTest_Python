# API 사용하여 crawling 해보기.
# API 인증 키. : hFryBABjlkn4SVTbLyCHYNoQSIOjSjcH%2B%2BsX74Bd0Ogr6vDe51JHzFq6t%2B7mIZtPBWtSaGQXjLeL4xROgx4zqg%3D%3D
# 키가 유출된다 해도 큰 문제가 없을 것 같다. 
# 사용 서비스 StatusOfPassengerFlights 여객운항 현황 조회
# Data type : XML
# 서비스 버전 1.0
# 사용 메소드 : getPassengerArrivals, getPassengerDepartures
# 도착
# 요청 메시지 명세
# from_time : 조회시간(부터), from_to : 조회시간(까지)
# airport 출발지 공항(공항 코드), flight_id 편명(KE130)
# airline 항공사 IATA(KE)
# lang : 언어구분(K)

# 응답 메시지 
# airline, flight_id, 
# scheduleDateTime, estimatedDateTime  도착 예정시간, 도착변경시간
# airport 출발 공항 gatenumber 탑승구 carousel 수하물 수취대번호
# exitnumber 출구(ABCD), remark 현황(도착, 결항, 지연, 회항, 착륙)
# airportCode 출발지 코드 IATA(KAL), terminalID (P01 제 1터미널)
# elapsetime 소요시간(1120 -> 11:20)
# firstopover 1경유지, firstopovername 1경유지 공항이름, secstopover : 2경유지. secstopovername 2경유지 이름
# thistopover 3경유지, thistopovername 3경유지 공항이름.

#출발
# 요청 메시지 명세
# from_time : 조회시간(부터), from_to : 조회시간(까지)
# airport 출발지 공항(공항 코드), flight_id 편명(KE130)
# airline 항공사 IATA(KE)
# lang : 언어구분(K)

# 응답 메시지 
# airline, flight_id, 
# scheduleDateTime, estimatedDateTime  출발 예정시간, 출발 변경시간
# airport 출발 공항, chkinrange 체크인 카운터, gatenumber 탑승구 carousel 수하물 수취대번호
# exitnumber 출구(ABCD), remark 현황(출발, 결항, 탑승중, 마감예정, 탑승마감, 탑승준비)
# airportCode 출발지 코드 IATA(KAL), terminalID (P01 제 1터미널)
# elapsetime 출발소요시간(1120 -> 11:20)
# firstopover 1경유지, firstopovername 1경유지 공항이름, secstopover : 2경유지. secstopovername 2경유지 이름
# thistopover 3경유지, thistopovername 3경유지 공항이름.

# 출발 조회, 도착 조회. 
# main 에서 키 인증

# Python 샘플 코드 #


import bs4
from urllib.request import urlopen
from urllib.parse import urlencode,unquote,quote_plus
import urllib
from lxml import html
import requests
import time
from openpyxl import Workbook, load_workbook


excel_file_path = './CodingTest_Python/Aircraft_position/'
excel_file_name = 'Aircraft_flight_' + str(time.strftime("%m%d", time.localtime(time.time()))) + ".xls"
# excel_sheet_title = str(time.strftime("%y-%m-%d, %H:%M", time.localtime(time.time())))
DEFAULT_FLIGHT_NAME = "KEXXX"
DEFAULT_DATE = "1 Sep 2009"
DEFAULT_TIME = "00:00"
DEFAULT_FLIGHT_STATUS = "NO_Flifhgt"
DEFAULT_LOCATION = "(ICN)"
EXCEL_START_ROW = 2
EXCEL_START_COL = 2

# http://openapi.airport.kr/openapi/service/StatusOfPassengerFlights/getPassengerDepartures?ServiceKey=hFryBABjlkn4SVTbLyCHYNoQSIOjSjcH%2B%2BsX74Bd0Ogr6vDe51JHzFq6t%2B7mIZtPBWtSaGQXjLeL4xROgx4zqg%3D%3D&airport_code=ICN
# http://openapi.airport.kr/openapi/service/StatusOfPassengerFlights/getPassengerDepartures?ServiceKey=hFryBABjlkn4SVTbLyCHYNoQSIOjSjcH%2B%2BsX74Bd0Ogr6vDe51JHzFq6t%2B7mIZtPBWtSaGQXjLeL4xROgx4zqg%3D%3D&airportCode=ICN
# http://openapi.airport.kr/openapi/service/StatusOfPassengerFlights/getPassengerDepartures?ServiceKey=hFryBABjlkn4SVTbLyCHYNoQSIOjSjcH%252B%252BsX74Bd0Ogr6vDe51JHzFq6t%252B7mIZtPBWtSaGQXjLeL4xROgx4zqg%253D%253D&airportCode=ICN
# print ("\nResult:")
# print (response_body)
# print ("\nDataType of Result Data:")
# print (type(response_body))
def set_Url(key, service_Type):
    url = url = 'http://openapi.airport.kr/openapi/service/' + service_Type
    queryParams = '?' + urlencode({quote_plus('ServiceKey') : key[:-1], quote_plus('airportCode') : 'ICN'})    

    #set api url
    request = urllib.request.Request(url+unquote(queryParams))
    print ('Your Request:\n'+url+queryParams)
    request.get_method = lambda: 'GET'
    with urlopen(request) as source :
        source = source.read()
        soup = bs4.BeautifulSoup(source, "html.parser")
        # print(soup.prettify())
    # //*[@id="collapsible4"]/div[1]/div[2]
    items = soup.find_all("item")
    return items

def get_flight_list(items):
    today_flight = []
    for item in items:
        flight_info = {}
        for element in item : 
            flight_info[element.name] = element.contents.pop()
            
        if(flight_info['airline'] == '대한항공'):
            # est_time = string_to_time(flight_info['estimateddatetime'])
            # sche_time = string_to_time(flight_info['estimateddatetime'])
            today_flight.append(flight_info)
            # print(flight_info)

    return today_flight


def insert_column(sheet1, columns, col_Value, flight_Info, row_Number):
    for i in range(len(columns)):
        sheet1.cell(row=EXCEL_START_ROW, column = EXCEL_START_COL+i).value = columns[i]
    # {'citycode': 'DFW', 'estimateddatetime': '0932', 'flightid': 'KE031', 'gatenumber': '231', 'remark': '출발', 'scheduledatetime': '0920', 'terminalid': 'P03'}
    # 시작 위치 (2,2)
    
    for _row in range(len(flight_Info)):
        for _col in range(len(col_Value)):
            if(flight_Info[_row].get(col_Value[_col])):
                sheet1.cell(row = EXCEL_START_ROW+_row+1, column = EXCEL_START_COL+_col, value=flight_Info[_row][col_Value[_col]])
                if(col_Value[_col] == 'terminalid') :
                    if( flight_Info[_row][col_Value[_col]] == 'P01' ):
                        sheet1.cell(row = EXCEL_START_ROW+_row+1, column = EXCEL_START_COL+_col, value='Terminal 1')
                    elif( flight_Info[_row][col_Value[_col]] == 'P02' ):
                        sheet1.cell(row = EXCEL_START_ROW+_row+1, column = EXCEL_START_COL+_col, value='Terminal 2')
                    else :
                        sheet1.cell(row = EXCEL_START_ROW+_row+1, column = EXCEL_START_COL+_col, value='Terminal 3')
            elif(col_Value[_col] == 'ICN'):
                sheet1.cell(row = EXCEL_START_ROW+_row+1, column = EXCEL_START_COL+_col, value='ICN')
            elif(col_Value[_col] == 'date'):
                sheet1.cell(row = EXCEL_START_ROW+_row+1, column = EXCEL_START_COL+_col, value=time.strftime("%y-%m-%d",time.localtime(time.time())))

def make_excel(depart_Flight_Info, arrival_Flight_Info, file_Name):
    # workbook 만들기
    
    wb =Workbook()
    # Sheet1 에 대한 옵션.
    sheet1 = wb.active
    sheet1.title = "Depart"
    # colume 값
    dep_Columns = ['Flight ID','Date','Flight Status', "Teminal Info", "GateNumber", "STD", "ETD","FROM","TO"]
    col_Depart_Value = ['flightid', 'date', 'remark', 'terminalid','gatenumber','scheduledatetime','estimateddatetime','ICN','airportcode']
    

    ari_Columns = ['Flight ID','Date','Flight Status', "Teminal Info", "GateNumber", "STA", "ETA","Flgiht_Time","FROM","TO"]
    col_Arrival_Value = ['flightid', 'date', 'remark', 'terminalid','gatenumber','scheduledatetime','estimateddatetime','elapsetime','airportcode','ICN']
    # {'airline': '대한항공', 'airportcode': 'LAX', 'citycode': 'LAX', 'elapsetime': '1216', 'estimateddatetime': '0438', 'flightid': 'KE012', 'gatenumber': '266', 'remark': '도착', 'scheduledatetime': '0500', 'terminalid': 'P03'}
   
    insert_column(sheet1, dep_Columns, col_Depart_Value, depart_Flight_Info, EXCEL_START_ROW)
    
    sheet2 = wb.create_sheet("Arrival",2)
    # sheet2 = wb.move_sheet(2)
    insert_column(sheet2, ari_Columns, col_Arrival_Value, arrival_Flight_Info, EXCEL_START_ROW)

    wb.save(filename=file_Name)
    wb.close()

def main():
    with open("./API_Key/key","rt") as f :
        key = f.readline()
        
    print(key)
    dep_Items = set_Url(key,'StatusOfPassengerFlights/getPassengerDepartures') # depart 부분. 
    ariv_Items  = set_Url(key,'StatusOfPassengerFlights/getPassengerArrivals') # depart 부분. 
    

    
    # get flighlist / dict 타입의 리스트
    depart_flight = get_flight_list(dep_Items)
    arrival_flight = get_flight_list(ariv_Items)
    make_excel(depart_flight,arrival_flight, excel_file_path+excel_file_name)
        # item:<item><airline>대한항공</airline><airport>두바이</airport><airportcode>DXB</airportcode><chkinrange>K19-K23</chkinrange><citycode>DXB</citycode><estimateddatetime>2355</estimateddatetime><flightid>KE5951</flightid><gatenumber>46</gatenumber><scheduledatetime>2355</scheduledatetime><terminalid>P01</terminalid></item>
        # 딕셔너리 데이터 이용하여 name 을 키, contents 를 value 로 하여 데이터 저장. 
        # 굳이 클래스 사용 필요 없

# def string_to_time(string_time):
#     hh = int(string_time[1]) + int(string_time[0])*10
#     mm = int(string_time[3]) + int(string_time[2])*10
#     return time.strftime('%h:%m', hh,mm)

        


main()
        
 


