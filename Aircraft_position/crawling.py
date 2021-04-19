import time
from datetime import datetime
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook, load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException


# [ 'HL8001','KE888', '09 May 2020', '2:39', 'Landed 08:17', 'STD', '—', 'ATD', '13:38', 'STA', '—', 'FROM', 'Xiamen (XMN)', 'TO', 'Seoul (ICN)']
class Flight(object):


    def __init__(self, flight_data ):
        self.flight_tail = flight_data[0]
        self.flight = flight_data[1]
        self.flight_time = flight_data[3]
        temp_status = flight_data[4].split(' ') # landed 00:00 estimated 00:00
        if len(temp_status)>1:
            # print(temp_status, flight_data[2])
            self.date = time_format(flight_data[2]+' '+temp_status[1])
        else :
            self.date = time_format(flight_data[2])
        self.flight_status = temp_status[0]
        self.atd = flight_data[8]
        self.loc_from = loc_setting(flight_data[12])
        self.loc_to = loc_setting(flight_data[14])

    def show_all(self):
        print("flight_tail",flight_tail,"flight_time : ",self.flight, "date : ",self.date,"flight_time : ", self.flight_time,"flight_status : ",self.flight_status,
        "flight_ATD : ", self.atd, "flight_loc_from : ",self.loc_from,"flight_loc_to : ", self.loc_to)

def loc_setting(str):
    # Seoul (ICN)
    try:
        if len(str)==0:
            str = '(ICN)'
        str1 = str.split('(')
        str2 = str1[1].split(')')
        loc = str2[0]
        return loc
    except Exception as e:
        print('wrong str', e)
        return 'ICN'

def time_format(str):
    try:
        # '09 May 2020'+'' 08:17'
        times = str.split(' ')
        if times:
            y=int(times[2])
            d=int(times[0])

            if times[1] == 'Jan': m = 1
            elif times[1] =='Feb': m = 2
            elif times[1] =='Mar': m = 3
            elif times[1] =='Apr': m = 4
            elif times[1] =='May': m = 5
            elif times[1] =='Jun': m = 6
            elif times[1] =='Jul': m = 7
            elif times[1] =='Aug': m = 8
            elif times[1] =='Sep': m = 9
            elif times[1] =='Oct': m = 10
            elif times[1] =='Nov': m = 11
            elif times[1] =='Dec': m =12
            if len(times)>2:
                t = times[3].split(':')
                hh = int(t[0])
                mm = int(t[1])
            else:
                hh = 0
                mm = 0
    except Exception as e:
        y = 2009
        m = 9
        d = 1
        hh = 0
        mm = 0
    return '{:%Y-%m-%d %H:%M}'.format(datetime(y, m, d, hh, mm))

def make_excel():

    wb =Workbook()
    sheet1 = wb.active
    sheet1.title = excel_sheet_title
    # colume 값
    #header
    columns = ['Tail','Flight_Name','Date','Flight_Duration','Flight Status','ATD','DST','ARI']
    for i in range(len(columns)):
        sheet1.cell(row=1, column = 1+i).value = columns[i]
    wb.save(filename=excel_file_name)
    wb.close()

# crawling results 에는 flights 리스트 가 들어있음.
def insert_data_to_excel(crawling_results):

    excel_file = load_workbook(excel_file_name)
    sheet1 = excel_file[excel_sheet_title]
    # columns = ['Flight_Name','Date','Flight_Duration','Flight Status','ATD','DST','ARI']
    excel_column= 1
    excel_row = 2

    for flight in crawling_results:
        columns = [ flight.flight_tail,flight.flight, flight.date, flight.flight_time, flight.flight_status,
        flight.atd, flight.loc_from, flight.loc_to ]
        # 각 cell 에 한개씩 저장, 한줄 저장 후 row 증가
        for i in range(len(columns)):
            sheet1.cell(row=excel_row, column=excel_column+i).value = columns[i]
        excel_row = excel_row +1

    excel_file.save(excel_file_name)
    excel_file.close()


if __name__ == '__main__':

    #https://ko.flightaware.com/live/flight/HL8001
    # https://www.flightradar24.com/data/aircraft/hl8001
    # 나중에 Tail 은 파일에서 가져올것.
    Base = 'https://www.flightradar24.com/data/aircraft/'
    Tails = ['HL7782','HL7783','HL7784','HL8208','HL8209','HL8210','HL8216','HL8217','HL8218','HL8250','HL8274','HL8275',
            'HL8211','HL8212','HL8227','HL8228','HL8276',
            'HL7532','HL7533','HL7534','HL7573',
            'HL7538','HL7539','HL7552',
            'HL7524','HL7525','HL7540','HL7550','HL7551','HL7553','HL7554','HL7584','HL7585','HL7586','HL7587','HL7702','HL7709','HL7710','HL7720',
            'HL7526','HL7530','HL7531','HL7574','HL7575','HL7598','HL7714','HL7715','HL7721',
            'HL8001','HL8002','HL8003','HL8025','HL8026','HL8027',
            'HL8081','HL8082','HL8083','HL8084','HL8085','HL7206','HL7207','HL7208','HL7209','HL8345']
    # Tails = ['HL8001','HL8002']


    excel_file_path = './'
    excel_file_name = 'Aircraft_flight.xlsx'
    excel_sheet_title = 'Today'
    # 창 없는 크롬 만들기.
    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    options.add_argument('window-size=960x1080') # 반 960 내 코드는 여기에서만 동작. 만약 큰 화면일경우 flight 함수 안에 들어가는 순서가 달라짐.
    options.add_argument('disable-gnu')
    # driver = webdriver.Chrome(ChromeDriverManager().install(), options = options)
    # driver = webdriver.Chrome(ChromeDriverManager().install())
    # /home/thales/Downloads/chromedriver_linux64
    driver = webdriver.Chrome(executable_path='/home/thales/Downloads/chromedriver_linux64/chromedriver', options = options)
    # driver = webdriver.Chrome(executable_path='/home/thales/Downloads/chromedriver_linux64/chromedriver')

    # driver = webdriver.Firefox(executable_path='/home/thales/Desktop/Thales_Tool/work/Python/geckodriver',firefox_options=options)

    flights = []
    for tail in Tails:
        try :
            url = Base + tail
            driver.get(url)
            time.sleep(1)
            element = WebDriverWait(driver,15).until(EC.presence_of_element_located((By.TAG_NAME,'tbody')))
            # html= driver.page_source
            # bs = BeautifulSoup(html,'html.parser')
            flight_data = []
            # xpath 를 통해서 검색하기.
            #tbody = driver.find_element_by_tag_name("tbody")
            tbody = driver.find_element_by_xpath('//*[@id="tbl-datatable"]/tbody')
            tr = tbody.find_element_by_class_name("data-row")
            if tr:
                flight_data.append(tail)
                # table 의 row 를 가져옴니다.
                # titles = tr.find_element_by_xpath('./*[@class="text-center-sm hidden-xs hidden-sm"]')
                # tds = tr.find_elements_by_class_name('hidden-xs hidden-sm')
                tds = tr.find_elements_by_tag_name('td')
                for td in tds:
                    if td.text:
                        td_list = []
                        td_list = td.text.split('\n')
                        flight_data.extend(td_list)
                #print(tail,flight_data)
                # for i in flight_data:
                #     print(i)
                flights.append(Flight(flight_data))
        except TimeoutException as e:
            print("time out exception,",e)
            driver.close()
            time.sleep(4)
            driver = webdriver.Chrome(executable_path='/home/thales/Downloads/chromedriver_linux64/chromedriver')
            driver.get(url)
        except NoSuchElementException as e :
            print(tail,"print_ref exception,",e)
            flight_data.extend([tail,'KEXXX','','0:0','NoFlight 00:00','','','','','','','','','(ICN)','(ICN)'])
            # ['KE888', '09 May 2020', '2:39', 'Landed 08:17', 'STD', '—', 'ATD', '13:38', 'STA', '—', 'FROM', 'Xiamen (XMN)', 'TO', 'Seoul (ICN)']
            #print(tail,flight_data)
            flights.append(Flight(flight_data))
        except IndexError as e:
            print(e,tail)

    driver.close()

    make_excel()

    insert_data_to_excel(flights)




# 완성
# ['KE888', ' KE888 ', ' 09 May 2020 ', ' 2:39 ', ' Landed 08:17 ']
# ['KE887', ' KE887 ', ' 09 May 2020 ', ' 2:28 ', ' Landed 04:02 ']
# ['KE888', ' KE888 ', ' 07 May 2020 ', ' 2:20 ', ' Landed 07:56 ']
# ['KE887', ' KE887 ', ' 07 May 2020 ', ' 2:35 ', ' Landed 04:01 ']
# ['KE888', ' KE888 ', ' 06 May 2020 ', ' 2:27 ', ' Landed 08:06 ']
# ['KE887', ' KE887 ', ' 06 May 2020 ', ' 2:34 ', ' Landed 04:14 ']
# ['KE888', ' KE888 ', ' 03 May 2020 ', ' 2:24 ', ' Landed 07:57 ']
# ['KE887', ' KE887 ', ' 03 May 2020 ', ' 2:32 ', ' Landed 04:03 ']
# trs = bs.find_all('div',{'class':'row table-row-responsive'}) # 한개의 데이터. 다른것 까지 찾을거면 find_all
# # divs1 = bs.find_all('div', {"class" : 'thrv_wrapper thrv_text_element'  })
# #flight_data = ['ke888','09 May 2020','2:39','Landed 08:17']
# # <div class="row"><i class="fa fa-plane"></i> <a class="fbold" href="/data/flights/ke888" target="_self">KE888</a> </div>
# # <div class="row"><i class="fa fa-calendar"></i> 09 May 2020 </div>
# # <div class="row"><i class="fa fa-clock-o"></i> 2:39 </div>
# # <div class="row"> Landed 08:17 </div>
#
# for tr in trs:
#     x3s = tr.find_all('div',{'class':'col-xs-3'})
#
#     for x3 in x3s:
#         rows = x3.find_all('div',{'class':'row'})
#         flight_data = []
#         for row in rows:
#             i_tag = row.find('i')
#             a_tag = row.find('a')
#             # flight_data = ['ke888','09 May 2020','2:39','Landed 08:17']
#             if a_tag:
#                 if a_tag.text:
#                     flight_data.append(a_tag.text)
#             if i_tag :
#                 if i_tag.text:
#                     flight_data.append(i_tag.text)
#             if row.text:
#                 flight_data.append(row.text)
#         print(flight_data)
